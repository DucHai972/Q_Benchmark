"""
Benchmark Pipeline for LLM Data Format Evaluation
"""

import os
import json
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Any, Optional
import logging
from dotenv import load_dotenv
import pandas as pd
from tqdm import tqdm

from benchmark_prompt_generator import BenchmarkPromptGenerator
from llm_clients import create_llm_client
from evaluator import BenchmarkEvaluator

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class BenchmarkPipeline:
    """Main benchmark pipeline for evaluating LLMs across data formats."""
    
    def __init__(self, output_dir: str = "benchmark_results", selected_models: Optional[List[str]] = None):
        """
        Initialize the benchmark pipeline.
        
        Args:
            output_dir: Directory to save results
            selected_models: List of models to initialize (openai, google, deepseek). If None, initialize all available.
        """
        load_dotenv()
        
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(exist_ok=True)
        self.selected_models = selected_models or ["openai", "google", "deepseek"]
        
        # Initialize components
        self.prompt_generator = BenchmarkPromptGenerator()
        self.evaluator = BenchmarkEvaluator()
        
        # Initialize LLM clients
        self.llm_clients = self._initialize_llm_clients()
        
        logger.info(f"Initialized pipeline with {len(self.llm_clients)} LLM clients")
    
    def _initialize_llm_clients(self) -> Dict[str, Any]:
        """Initialize LLM clients from environment variables for selected models only."""
        clients = {}
        
        # OpenAI
        if "openai" in self.selected_models:
            openai_key = os.getenv("OPENAI_API_KEY")
            if openai_key:
                try:
                    clients["openai"] = create_llm_client("openai", openai_key)
                    logger.info("Initialized OpenAI client")
                except Exception as e:
                    logger.warning(f"Failed to initialize OpenAI client: {e}")
            else:
                logger.warning("OpenAI selected but OPENAI_API_KEY not found in .env")
        
        # Google
        if "google" in self.selected_models:
            google_key = os.getenv("GOOGLE_API_KEY")
            if google_key:
                try:
                    clients["google"] = create_llm_client("google", google_key)
                    logger.info("Initialized Google client")
                except Exception as e:
                    logger.warning(f"Failed to initialize Google client: {e}")
            else:
                logger.warning("Google selected but GOOGLE_API_KEY not found in .env")
        
        # DeepSeek
        if "deepseek" in self.selected_models:
            deepseek_key = os.getenv("DEEPSEEK_API_KEY")
            if deepseek_key:
                try:
                    clients["deepseek"] = create_llm_client("deepseek", deepseek_key)
                    logger.info("Initialized DeepSeek client")
                except Exception as e:
                    logger.warning(f"Failed to initialize DeepSeek client: {e}")
            else:
                logger.warning("DeepSeek selected but DEEPSEEK_API_KEY not found in .env")
        
        if not clients:
            raise ValueError(f"No valid API keys found for selected models: {self.selected_models}. Please check your .env file.")
        
        return clients
    
    def run_benchmark(self, 
                     dataset: Optional[str] = None,
                     task: Optional[str] = None, 
                     max_cases: Optional[int] = None,
                     formats: Optional[List[str]] = None) -> str:
        """
        Run the complete benchmark pipeline.
        
        Args:
            dataset: Specific dataset to test (None for all)
            task: Specific task to test (None for all)
            max_cases: Maximum cases per task/dataset
            formats: Specific formats to test (None for all)
            
        Returns:
            Path to results directory
        """
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        run_dir = self.output_dir / f"run_{timestamp}"
        run_dir.mkdir(exist_ok=True)
        
        return self._run_benchmark_core(run_dir, dataset, task, max_cases, formats)
    
    def run_benchmark_fixed_dir(self, 
                                run_dir: str,
                                dataset: Optional[str] = None,
                                task: Optional[str] = None, 
                                max_cases: Optional[int] = None,
                                formats: Optional[List[str]] = None,
                                progress_bar=None) -> str:
        """
        Run benchmark using a fixed run directory.
        
        Args:
            run_dir: Fixed run directory path
            dataset: Specific dataset to test (None for all)
            task: Specific task to test (None for all)
            max_cases: Maximum cases per task/dataset
            formats: Specific formats to test (None for all)
            
        Returns:
            Path to results directory
        """
        run_dir_path = Path(run_dir)
        # Ensure the run directory exists
        run_dir_path.mkdir(parents=True, exist_ok=True)
        return self._run_benchmark_core(run_dir_path, dataset, task, max_cases, formats, progress_bar)
    
    def _run_benchmark_core(self, 
                           run_dir: Path,
                           dataset: Optional[str] = None,
                           task: Optional[str] = None, 
                           max_cases: Optional[int] = None,
                           formats: Optional[List[str]] = None,
                           progress_bar=None) -> str:
        """
        Core benchmark logic.
        
        Returns:
            Path to results directory
        """
        
        logger.info(f"Starting benchmark run: {run_dir}")
        
        # Generate prompts
        logger.info("Generating benchmark prompts...")
        prompts = self.prompt_generator.generate_benchmark_suite(
            dataset=dataset, 
            task=task, 
            max_cases=max_cases
        )
        
        # Filter formats if specified
        if formats:
            prompts = [p for p in prompts if p['data_format'] in formats]
        
        logger.info(f"Generated {len(prompts)} prompts")
        
        # Group prompts by dataset and task
        grouped_prompts = {}
        for prompt in prompts:
            key = f"{prompt['dataset']}_{prompt['task']}"
            if key not in grouped_prompts:
                grouped_prompts[key] = []
            grouped_prompts[key].append(prompt)
        
        # Run evaluation on each LLM for each dataset/task group
        total_model_evaluations = len(grouped_prompts) * len(self.llm_clients)
        current_evaluation = 0
        
        for group_key, group_prompts in grouped_prompts.items():
            dataset_name, task_name = group_key.split('_', 1)
            group_dir = run_dir / dataset_name / task_name
            group_dir.mkdir(parents=True, exist_ok=True)
            
            for provider_name, client in self.llm_clients.items():
                current_evaluation += 1
                logger.info(f"Evaluating {provider_name} on {group_key}... ({current_evaluation}/{total_model_evaluations})")
                
                # Generate simple results
                simple_results = self._evaluate_llm_simple(client, group_prompts, progress_bar)
                
                # Save results with model name
                model_name_safe = client.model_name.replace("/", "_").replace("-", "_")
                results_file = group_dir / f"{model_name_safe}_results.json"
                with open(results_file, 'w', encoding='utf-8') as f:
                    json.dump(simple_results, f, indent=2, ensure_ascii=False)
        
        logger.info(f"Benchmark completed. Results saved to: {run_dir}")
        return str(run_dir)
    
    def _evaluate_llm_simple(self, client, prompts: List[Dict[str, Any]], progress_bar=None) -> List[Dict[str, Any]]:
        """Evaluate LLM and return simplified results."""
        results = []
        
        for i, prompt_data in enumerate(prompts):
            # Generate LLM response
            llm_result = client.generate(prompt_data['prompt'])
            
            # Evaluate response
            score = 0.0
            is_correct = False
            if llm_result['success'] and llm_result['response']:
                evaluation = self.evaluator.evaluate_response(
                    llm_result['response'],
                    prompt_data['expected_answer'],
                    prompt_data['task']
                )
                score = evaluation['score']
                is_correct = score > 0.8
            
            # Create simple result
            result = {
                "case_id": prompt_data['case_id'],
                "data_format": prompt_data['data_format'],
                "model_name": client.model_name,
                "prompt": prompt_data['prompt'],
                "expected_answer": prompt_data['expected_answer'],
                "llm_response": llm_result.get('response', ''),
                "is_correct": is_correct,
                "score": score,
                "response_time": llm_result.get('response_time', 0),
                "success": llm_result.get('success', False),
                "error": llm_result.get('error', None)
            }
            
            results.append(result)
            
            # Update progress bar if provided
            if progress_bar:
                progress_bar.update(1)
            
            # Add delay to avoid rate limits
            time.sleep(0.5)
        
        return results
    
    def _generate_simple_summary(self, run_dir: Path):
        """Generate a simple summary of all results."""
        summary = {
            "timestamp": datetime.now().isoformat(),
            "overall_stats": {},
            "by_provider": {},
            "by_format": {}
        }
        
        all_results = []
        
        # Collect all results
        for dataset_dir in run_dir.iterdir():
            if not dataset_dir.is_dir():
                continue
            for task_dir in dataset_dir.iterdir():
                if not task_dir.is_dir():
                    continue
                for result_file in task_dir.glob("*_results.json"):
                    with open(result_file, 'r') as f:
                        data = json.load(f)
                    
                    for item in data:
                        # Use model_name from the data instead of filename
                        item['provider'] = item.get('model_name', 'unknown')
                        item['dataset'] = dataset_dir.name
                        item['task'] = task_dir.name
                        all_results.append(item)
        
        # Calculate statistics
        if all_results:
            total_cases = len(all_results)
            successful_cases = sum(1 for r in all_results if r['success'])
            correct_cases = sum(1 for r in all_results if r['is_correct'])
            
            summary["overall_stats"] = {
                "total_cases": total_cases,
                "success_rate": successful_cases / total_cases if total_cases > 0 else 0,
                "accuracy_rate": correct_cases / successful_cases if successful_cases > 0 else 0,
                "avg_score": sum(r['score'] for r in all_results) / len(all_results)
            }
            
            # By provider
            providers = set(r['provider'] for r in all_results)
            for provider in providers:
                provider_results = [r for r in all_results if r['provider'] == provider]
                successful = sum(1 for r in provider_results if r['success'])
                correct = sum(1 for r in provider_results if r['is_correct'])
                
                summary["by_provider"][provider] = {
                    "total_cases": len(provider_results),
                    "success_rate": successful / len(provider_results),
                    "accuracy_rate": correct / successful if successful > 0 else 0,
                    "avg_score": sum(r['score'] for r in provider_results) / len(provider_results)
                }
            
            # By format
            formats = set(r['data_format'] for r in all_results)
            for data_format in formats:
                format_results = [r for r in all_results if r['data_format'] == data_format]
                successful = sum(1 for r in format_results if r['success'])
                correct = sum(1 for r in format_results if r['is_correct'])
                
                summary["by_format"][data_format] = {
                    "total_cases": len(format_results),
                    "success_rate": successful / len(format_results),
                    "accuracy_rate": correct / successful if successful > 0 else 0,
                    "avg_score": sum(r['score'] for r in format_results) / len(format_results)
                }
        
        # Save summary
        summary_file = run_dir / "summary.json"
        with open(summary_file, 'w', encoding='utf-8') as f:
            json.dump(summary, f, indent=2, ensure_ascii=False)
    
    def _generate_excel_table(self, run_dir: Path):
        """Generate Excel table with data formats vs benchmark tasks/LLM models."""
        logger.info("Generating Excel results table...")
        
        all_results = []
        
        # Collect all results
        for dataset_dir in run_dir.iterdir():
            if not dataset_dir.is_dir():
                continue
            for task_dir in dataset_dir.iterdir():
                if not task_dir.is_dir():
                    continue
                for result_file in task_dir.glob("*_results.json"):
                    with open(result_file, 'r') as f:
                        data = json.load(f)
                    
                    for item in data:
                        # Use model_name from the data instead of filename
                        item['provider'] = item.get('model_name', 'unknown')
                        item['dataset'] = dataset_dir.name
                        item['task'] = task_dir.name
                        all_results.append(item)
        
        if not all_results:
            logger.warning("No results found for Excel generation")
            return
        
        # Get unique values
        formats = sorted(set(r['data_format'] for r in all_results))
        datasets = sorted(set(r['dataset'] for r in all_results))
        tasks = sorted(set(r['task'] for r in all_results))
        providers = sorted(set(r['provider'] for r in all_results))
        
        # Create multi-level columns: (dataset, task, provider)
        columns = []
        for dataset in datasets:
            for task in tasks:
                for provider in providers:
                    # Check if this combination exists in results
                    combo_exists = any(
                        r['dataset'] == dataset and r['task'] == task and r['provider'] == provider 
                        for r in all_results
                    )
                    if combo_exists:
                        columns.append((dataset, task, provider))
        
        # Create DataFrame
        data_matrix = []
        
        for data_format in formats:
            row = [data_format]  # First column is data format
            
            for dataset, task, provider in columns:
                # Find results for this combination
                format_results = [
                    r for r in all_results 
                    if (r['data_format'] == data_format and 
                        r['dataset'] == dataset and 
                        r['task'] == task and 
                        r['provider'] == provider)
                ]
                
                if format_results:
                    total_cases = len(format_results)
                    correct_cases = sum(1 for r in format_results if r['is_correct'])
                    cell_value = f"{correct_cases}/{total_cases}"
                else:
                    cell_value = "0/0"
                
                row.append(cell_value)
            
            data_matrix.append(row)
        
        # Create column headers
        column_headers = ['Data Format']
        for dataset, task, provider in columns:
            column_headers.append(f"{dataset}_{task}_{provider}")
        
        # Create DataFrame
        df = pd.DataFrame(data_matrix)
        df.columns = column_headers
        
        # Create Excel file with multi-level headers
        excel_file = run_dir / "benchmark_results_table.xlsx"
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Write the main data
            df.to_excel(writer, sheet_name='Results', index=False, startrow=2)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Results']
            
            # Create multi-level headers
            # Row 1: Dataset names
            # Row 2: Task names  
            # Row 3: Provider names
            
            col_idx = 2  # Start from column B (after Data Format)
            for dataset, task, provider in columns:
                # Write dataset name in row 1
                worksheet.cell(row=1, column=col_idx, value=dataset)
                # Write task name in row 2
                worksheet.cell(row=2, column=col_idx, value=task)
                # Write provider name in row 3  
                worksheet.cell(row=3, column=col_idx, value=provider)
                col_idx += 1
            
            # Write "Data Format" in the first column
            worksheet.cell(row=1, column=1, value="Data")
            worksheet.cell(row=2, column=1, value="Format")
            worksheet.cell(row=3, column=1, value="")
            
            # Merge cells for repeated dataset/task names
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment, Border, Side
            
            # Add borders and center alignment
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'), 
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply formatting to all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Merge dataset and task cells where they repeat
            current_dataset = None
            current_task = None
            dataset_start_col = 2
            task_start_col = 2
            
            col_idx = 2
            for dataset, task, provider in columns:
                if dataset != current_dataset:
                    # Merge previous dataset cells if needed
                    if current_dataset and col_idx > dataset_start_col:
                        if col_idx - 1 > dataset_start_col:
                            worksheet.merge_cells(
                                start_row=1, start_column=dataset_start_col,
                                end_row=1, end_column=col_idx - 1
                            )
                    current_dataset = dataset
                    dataset_start_col = col_idx
                
                if task != current_task or dataset != current_dataset:
                    # Merge previous task cells if needed
                    if current_task and col_idx > task_start_col:
                        if col_idx - 1 > task_start_col:
                            worksheet.merge_cells(
                                start_row=2, start_column=task_start_col,
                                end_row=2, end_column=col_idx - 1
                            )
                    current_task = task
                    task_start_col = col_idx
                
                col_idx += 1
            
            # Merge final dataset and task cells
            if col_idx > dataset_start_col:
                if col_idx - 1 > dataset_start_col:
                    worksheet.merge_cells(
                        start_row=1, start_column=dataset_start_col,
                        end_row=1, end_column=col_idx - 1
                    )
            
            if col_idx > task_start_col:
                if col_idx - 1 > task_start_col:
                    worksheet.merge_cells(
                        start_row=2, start_column=task_start_col,
                        end_row=2, end_column=col_idx - 1
                    )
            
            # Merge Data Format header
            worksheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
        
        logger.info(f"Excel table saved to: {excel_file}")
    
    def _generate_score_excel_table(self, run_dir: Path):
        """Generate Excel table with score values (score/max_score) instead of correct/incorrect counts."""
        logger.info("Generating score-based Excel results table...")
        
        all_results = []
        
        # Collect all results
        for dataset_dir in run_dir.iterdir():
            if not dataset_dir.is_dir():
                continue
            for task_dir in dataset_dir.iterdir():
                if not task_dir.is_dir():
                    continue
                for result_file in task_dir.glob("*_results.json"):
                    with open(result_file, 'r') as f:
                        data = json.load(f)
                    
                    for item in data:
                        # Use model_name from the data instead of filename
                        item['provider'] = item.get('model_name', 'unknown')
                        item['dataset'] = dataset_dir.name
                        item['task'] = task_dir.name
                        all_results.append(item)
        
        if not all_results:
            logger.warning("No results found for score Excel generation")
            return
        
        # Get unique values
        formats = sorted(set(r['data_format'] for r in all_results))
        datasets = sorted(set(r['dataset'] for r in all_results))
        tasks = sorted(set(r['task'] for r in all_results))
        providers = sorted(set(r['provider'] for r in all_results))
        
        # Create multi-level columns: (dataset, task, provider)
        columns = []
        for dataset in datasets:
            for task in tasks:
                for provider in providers:
                    # Check if this combination exists in results
                    combo_exists = any(
                        r['dataset'] == dataset and r['task'] == task and r['provider'] == provider 
                        for r in all_results
                    )
                    if combo_exists:
                        columns.append((dataset, task, provider))
        
        # Create DataFrame
        data_matrix = []
        
        for data_format in formats:
            row = [data_format]  # First column is data format
            
            for dataset, task, provider in columns:
                # Find results for this combination
                format_results = [
                    r for r in all_results 
                    if (r['data_format'] == data_format and 
                        r['dataset'] == dataset and 
                        r['task'] == task and 
                        r['provider'] == provider)
                ]
                
                if format_results:
                    total_cases = len(format_results)
                    total_score = sum(r['score'] for r in format_results)
                    avg_score = total_score / total_cases if total_cases > 0 else 0
                    # Format as "total_score/total_cases" (e.g., "8.5/10")
                    cell_value = f"{total_score:.1f}/{total_cases}"
                else:
                    cell_value = "0.0/0"
                
                row.append(cell_value)
            
            data_matrix.append(row)
        
        # Create column headers
        column_headers = ['Data Format']
        for dataset, task, provider in columns:
            column_headers.append(f"{dataset}_{task}_{provider}")
        
        # Create DataFrame
        df = pd.DataFrame(data_matrix)
        df.columns = column_headers
        
        # Create Excel file with multi-level headers
        excel_file = run_dir / "benchmark_scores_table.xlsx"
        
        with pd.ExcelWriter(excel_file, engine='openpyxl') as writer:
            # Write the main data
            df.to_excel(writer, sheet_name='Scores', index=False, startrow=2)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Scores']
            
            # Create multi-level headers
            # Row 1: Dataset names
            # Row 2: Task names  
            # Row 3: Provider names
            
            col_idx = 2  # Start from column B (after Data Format)
            for dataset, task, provider in columns:
                # Write dataset name in row 1
                worksheet.cell(row=1, column=col_idx, value=dataset)
                # Write task name in row 2
                worksheet.cell(row=2, column=col_idx, value=task)
                # Write provider name in row 3  
                worksheet.cell(row=3, column=col_idx, value=provider)
                col_idx += 1
            
            # Write "Data Format" in the first column
            worksheet.cell(row=1, column=1, value="Data")
            worksheet.cell(row=2, column=1, value="Format")
            worksheet.cell(row=3, column=1, value="")
            
            # Merge cells for repeated dataset/task names
            from openpyxl.utils import get_column_letter
            from openpyxl.styles import Alignment, Border, Side
            
            # Add borders and center alignment
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'), 
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Apply formatting to all cells
            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border
                    cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Merge dataset and task cells where they repeat
            current_dataset = None
            current_task = None
            dataset_start_col = 2
            task_start_col = 2
            
            col_idx = 2
            for dataset, task, provider in columns:
                if dataset != current_dataset:
                    # Merge previous dataset cells if needed
                    if current_dataset and col_idx > dataset_start_col:
                        if col_idx - 1 > dataset_start_col:
                            worksheet.merge_cells(
                                start_row=1, start_column=dataset_start_col,
                                end_row=1, end_column=col_idx - 1
                            )
                    current_dataset = dataset
                    dataset_start_col = col_idx
                
                if task != current_task or dataset != current_dataset:
                    # Merge previous task cells if needed
                    if current_task and col_idx > task_start_col:
                        if col_idx - 1 > task_start_col:
                            worksheet.merge_cells(
                                start_row=2, start_column=task_start_col,
                                end_row=2, end_column=col_idx - 1
                            )
                    current_task = task
                    task_start_col = col_idx
                
                col_idx += 1
            
            # Merge final dataset and task cells
            if col_idx > dataset_start_col:
                if col_idx - 1 > dataset_start_col:
                    worksheet.merge_cells(
                        start_row=1, start_column=dataset_start_col,
                        end_row=1, end_column=col_idx - 1
                    )
            
            if col_idx > task_start_col:
                if col_idx - 1 > task_start_col:
                    worksheet.merge_cells(
                        start_row=2, start_column=task_start_col,
                        end_row=2, end_column=col_idx - 1
                    )
            
            # Merge Data Format header
            worksheet.merge_cells(start_row=1, start_column=1, end_row=3, end_column=1)
        
        logger.info(f"Score Excel table saved to: {excel_file}")

def main():
    """Run benchmark with default settings."""
    pipeline = BenchmarkPipeline()
    
    # Run benchmark on a subset for testing
    results_dir = pipeline.run_benchmark(
        dataset="healthcare-dataset",  # Focus on one dataset
        task="answer_lookup",          # Focus on one task
        max_cases=5                    # Limit cases for testing
    )
    
    print(f"Benchmark completed. Results in: {results_dir}")

if __name__ == "__main__":
    main() 